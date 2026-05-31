#!/usr/bin/env python3
"""
Export Attendance + Attendance_2568 → SQL INSERT statements
รันใน Supabase SQL Editor ได้เลย
"""
import openpyxl
from datetime import datetime

SCHOOL_CODE = 'BK001'
SOURCE = 'สารสนเทศ_บ้านแก่ง.xlsx'
OUTPUT = 'import_attendance.sql'

wb = openpyxl.load_workbook(SOURCE, data_only=True)

all_rows = []
for sheet_name in ['Attendance', 'Attendance_2568']:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f'📖 Reading "{sheet_name}" ({ws.max_row - 1} rows)...')
    for row in range(2, ws.max_row + 1):
        date  = ws.cell(row, 2).value
        sid   = ws.cell(row, 3).value
        cls   = ws.cell(row, 4).value
        stat  = ws.cell(row, 5).value
        rem   = ws.cell(row, 6).value

        if not (date and sid and stat):
            continue

        # Date → ISO
        if isinstance(date, datetime):
            date_iso = date.strftime('%Y-%m-%d')
        else:
            date_iso = str(date)[:10]

        # StudentID → string (handle 4590.0 → "4590")
        sid_str = str(int(sid)) if isinstance(sid, float) else str(sid).strip()
        cls_str = str(cls).strip() if cls else ''
        stat_str = str(stat).strip()
        rem_str = str(rem).strip() if rem else ''

        # Escape single quotes for SQL
        def esc(s): return s.replace("'", "''")
        cls_sql = f"'{esc(cls_str)}'" if cls_str else 'NULL'
        rem_sql = f"'{esc(rem_str)}'" if rem_str else 'NULL'

        all_rows.append(
            f"  ('{date_iso}'::date, '{sid_str}', {cls_sql}, '{esc(stat_str)}', {rem_sql})"
        )

print(f'\n📊 Total: {len(all_rows)} attendance records')

# write SQL file
with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write('-- ============================================\n')
    f.write(f'-- Import Attendance ({len(all_rows)} records)\n')
    f.write('-- ============================================\n\n')

    # split into batches of 500 to keep transactions manageable
    BATCH = 500
    for i in range(0, len(all_rows), BATCH):
        chunk = all_rows[i:i + BATCH]
        f.write(f'-- Batch {i // BATCH + 1}: rows {i + 1} to {i + len(chunk)}\n')
        f.write(
            'with sch as (select id from schools where code = \'' + SCHOOL_CODE + '\')\n'
            'insert into attendance (school_id, attendance_date, student_id, class, status, remark)\n'
            'select sch.id, t.attendance_date, t.student_id, t.class, t.status, t.remark\n'
            'from sch, (values\n'
        )
        f.write(',\n'.join(chunk))
        f.write('\n) as t(attendance_date, student_id, class, status, remark)\n')
        f.write('on conflict (school_id, attendance_date, student_id) do update set\n')
        f.write('  status = excluded.status, class = excluded.class, remark = excluded.remark;\n\n')

    f.write(f"-- ตรวจสอบ\nselect count(*) from attendance;\n")

print(f'✅ เขียน {OUTPUT} เสร็จ')
print(f'   → เปิดไฟล์ → copy → paste ใน Supabase SQL Editor → Run')
