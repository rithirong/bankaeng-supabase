#!/usr/bin/env python3
"""
Export Savings + Coop จาก Google Sheets export → SQL INSERT
ใช้ NOT EXISTS เพื่อ skip student ที่ไม่มีในระบบ
"""
import openpyxl
from datetime import datetime

SCHOOL_CODE = 'BK001'
SOURCE = 'สารสนเทศ_บ้านแก่ง.xlsx'
OUTPUT = 'import_savings_coop.sql'

def get_academic_year(d):
    """ปีการศึกษา: เริ่ม 16 พ.ค. ของแต่ละปี"""
    if not d: return None
    y = d.year + 543
    return y if d.month >= 5 else y - 1

def esc(s):
    return str(s).replace("'", "''")

wb = openpyxl.load_workbook(SOURCE, data_only=True)

# ────────────────────────────────────────
# 💰 SAVINGS
# ────────────────────────────────────────
savings_rows = []
for sheet_name in ['Savings', 'Savings_2569', 'Savings_2568']:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f'📖 Reading "{sheet_name}" ({ws.max_row - 1} rows)...')
    for row in range(2, ws.max_row + 1):
        date  = ws.cell(row, 2).value
        sid   = ws.cell(row, 3).value
        cls   = ws.cell(row, 4).value
        typ   = ws.cell(row, 5).value
        amt   = ws.cell(row, 6).value

        if not (date and sid and typ and amt):
            continue
        if isinstance(date, datetime):
            date_iso = date.strftime('%Y-%m-%d')
        else:
            date_iso = str(date)[:10]
        sid_str = str(int(sid)) if isinstance(sid, float) else str(sid).strip()
        cls_str = str(cls).strip() if cls else ''
        typ_str = str(typ).strip()
        if typ_str not in ('ฝาก', 'ถอน'):
            continue
        try:
            amt_val = float(amt)
        except:
            continue
        if amt_val <= 0:
            continue
        year = get_academic_year(date if isinstance(date, datetime) else datetime.strptime(date_iso, '%Y-%m-%d'))

        cls_sql = f"'{esc(cls_str)}'" if cls_str else 'NULL'
        savings_rows.append(
            f"  ('{date_iso}'::date, {year}, '{sid_str}', {cls_sql}, '{esc(typ_str)}', {amt_val})"
        )

print(f'  → {len(savings_rows)} savings records (skip dup ใน DB ด้วย NOT EXISTS)')

# ────────────────────────────────────────
# 🛒 COOP (สหกรณ์ร้านค้า)
# ────────────────────────────────────────
coop_rows = []
if 'สหกรณ์ร้านค้า' in wb.sheetnames:
    ws = wb['สหกรณ์ร้านค้า']
    print(f'📖 Reading "สหกรณ์ร้านค้า" ({ws.max_row - 1} rows)...')
    for row in range(2, ws.max_row + 1):
        date  = ws.cell(row, 1).value
        typ   = ws.cell(row, 2).value
        cat   = ws.cell(row, 3).value
        desc  = ws.cell(row, 4).value
        amt   = ws.cell(row, 5).value
        url   = ws.cell(row, 6).value

        if not (date and typ and amt):
            continue
        if isinstance(date, datetime):
            date_iso = date.strftime('%Y-%m-%d')
        else:
            date_iso = str(date)[:10]
        typ_str = str(typ).strip()
        if typ_str not in ('รายรับ', 'รายจ่าย'):
            continue
        try:
            amt_val = float(amt)
        except:
            continue
        if amt_val <= 0:
            continue
        cat_sql  = f"'{esc(str(cat).strip())}'" if cat else 'NULL'
        desc_sql = f"'{esc(str(desc).strip())}'" if desc else 'NULL'
        url_sql  = f"'{esc(str(url).strip())}'" if url else 'NULL'
        coop_rows.append(
            f"  ('{date_iso}'::date, '{esc(typ_str)}', {cat_sql}, {desc_sql}, {amt_val}, {url_sql})"
        )

print(f'  → {len(coop_rows)} coop entries')

# ────────────────────────────────────────
# Write SQL
# ────────────────────────────────────────
with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write('-- ============================================\n')
    f.write(f'-- Import Savings ({len(savings_rows)}) + Coop ({len(coop_rows)})\n')
    f.write('-- ============================================\n\n')

    # Clear old
    f.write('-- ล้างก่อน import (กัน duplicate)\n')
    f.write(f"delete from savings where school_id = (select id from schools where code = '{SCHOOL_CODE}');\n")
    f.write(f"delete from coop_entries where school_id = (select id from schools where code = '{SCHOOL_CODE}');\n\n")

    # SAVINGS — batch 500
    if savings_rows:
        BATCH = 500
        for i in range(0, len(savings_rows), BATCH):
            chunk = savings_rows[i:i + BATCH]
            f.write(f'-- Savings batch {i // BATCH + 1}: rows {i + 1} to {i + len(chunk)}\n')
            f.write(
                f"with sch as (select id from schools where code = '{SCHOOL_CODE}')\n"
                'insert into savings (school_id, txn_date, academic_year, student_id, class, type, amount)\n'
                'select sch.id, t.txn_date, t.academic_year, t.student_id, t.class, t.type, t.amount\n'
                'from sch, (values\n'
            )
            f.write(',\n'.join(chunk))
            f.write('\n) as t(txn_date, academic_year, student_id, class, type, amount)\n')
            # filter: skip student ที่ไม่มีในระบบ
            f.write("where exists (select 1 from students s where s.school_id = sch.id and s.student_id = t.student_id);\n\n")

    # COOP
    if coop_rows:
        f.write(f'-- Coop entries\n')
        f.write(
            f"with sch as (select id from schools where code = '{SCHOOL_CODE}')\n"
            'insert into coop_entries (school_id, entry_date, type, category, description, amount, receipt_url)\n'
            'select sch.id, t.entry_date, t.type, t.category, t.description, t.amount, t.receipt_url\n'
            'from sch, (values\n'
        )
        f.write(',\n'.join(coop_rows))
        f.write('\n) as t(entry_date, type, category, description, amount, receipt_url);\n\n')

    f.write('-- ตรวจสอบ\n')
    f.write("select 'savings' as t, count(*) from savings\n")
    f.write("union all select 'coop_entries', count(*) from coop_entries;\n")

print(f'\n✅ เขียน {OUTPUT} เสร็จ')
print(f'   → เปิดไฟล์ → copy → paste ใน Supabase SQL Editor → Run')
