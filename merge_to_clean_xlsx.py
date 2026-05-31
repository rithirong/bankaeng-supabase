#!/usr/bin/env python3
"""
Merge DMC + Sheets export → ไฟล์ .xlsx สะอาดไฟล์เดียว
ใช้กลยุทธ์:
- standard fields (A-AJ): ใช้ DMC เป็นหลัก (address เป็น string สะอาด)
- AK-AO: ใช้ Sheets (status, current_academic_year)
- นักเรียนที่อยู่ใน Sheets แต่ไม่อยู่ DMC: ใช้ Sheets ทั้งหมด + กู้ address จาก Date
- Output: clean.xlsx → upload เข้า /admin/import ได้เลย
"""

import openpyxl
from datetime import datetime
from openpyxl import Workbook

DMC_FILE = 'ข้อมูลนร. 10 มิ.ย68.xlsx'
SHEETS_FILE = 'สารสนเทศ_บ้านแก่ง.xlsx'
OUTPUT_FILE = 'clean_students.xlsx'

# ── อ่าน DMC (row 1 = metadata, row 2 = headers, row 3+ = data) ──
def read_dmc():
    wb = openpyxl.load_workbook(DMC_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    data = {}
    for row in range(3, ws.max_row + 1):
        sid = ws.cell(row=row, column=6).value  # F = student_id
        if not sid: continue
        sid = str(int(sid)) if isinstance(sid, float) else str(sid).strip()
        if not sid: continue
        rec = {}
        for c, h in enumerate(headers, start=1):
            rec[h] = ws.cell(row=row, column=c).value
        data[sid] = rec
    return data

# ── อ่าน Sheets (row 1 = headers, row 2+ = data) ──
def read_sheets():
    wb = openpyxl.load_workbook(SHEETS_FILE, data_only=True)
    ws = wb['Student']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # AK header "ปกติ" จริงๆ คือ status; AO header "2568"/year จริงๆ คือ current_academic_year
    headers[36] = 'สถานะ'  # AK (index 36) override → status
    headers[40] = 'ปีการศึกษาปัจจุบัน'  # AO (index 40) override → year
    data = {}
    for row in range(2, ws.max_row + 1):
        sid = ws.cell(row=row, column=6).value
        if not sid: continue
        sid = str(int(sid)) if isinstance(sid, float) else str(sid).strip()
        if not sid: continue
        rec = {}
        for c, h in enumerate(headers, start=1):
            rec[h] = ws.cell(row=row, column=c).value
        data[sid] = rec
    return data

# ── ฟื้น address ที่เป็น Date กลับเป็น "D/M" ──
def fix_address(val):
    if isinstance(val, datetime):
        return f'{val.day}/{val.month}'
    return val

# ── แปลง value เป็น string สะอาด ──
def to_str(v):
    if v is None: return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()

# ── Merge ──
print('📂 อ่าน DMC...')
dmc = read_dmc()
print(f'   → {len(dmc)} นักเรียน')

print('📂 อ่าน Sheets...')
sheets = read_sheets()
print(f'   → {len(sheets)} นักเรียน')

# union ของ student_id ทั้งหมด
all_sids = sorted(set(dmc.keys()) | set(sheets.keys()), key=lambda x: int(x) if x.isdigit() else 99999)

# headers สำหรับ output (ใช้ headers ของ Sheets เป็นโครง + override fixes)
output_headers = [
    'รหัสโรงเรียน', 'ชื่อโรงเรียน',
    'เลขประจำตัวประชาชน',  # was C "เลขประจำตัวนักเรียน" 13-digit
    'ชั้น', 'ห้อง',
    'เลขประจำตัวนักเรียน',  # F 4-digit
    'เพศ', 'คำนำหน้าชื่อ', 'ชื่อ', 'นามสกุล',
    'วันเกิด', 'อายุ(ปี)', 'น้ำหนัก', 'ส่วนสูง', 'กลุ่มเลือด',
    'ศาสนา', 'เชื้อชาติ', 'สัญชาติ',
    'บ้านเลขที่', 'หมู่', 'ถนน/ซอย', 'ตำบล', 'อำเภอ', 'จังหวัด',
    'ชื่อผู้ปกครอง', 'นามสกุลผู้ปกครอง', 'อาชีพของผู้ปกครอง',
    'ความเกี่ยวข้องของผู้ปกครองกับนักเรียน',
    'ชื่อบิดา', 'นามสกุลบิดา', 'อาชีพของบิดา',
    'ชื่อมารดา', 'นามสกุลมารดา', 'อาชีพของมารดา',
    'ความด้อยโอกาส',
    'สถานะ',  # from Sheets AK
    'ปีการศึกษาปัจจุบัน',  # from Sheets AO
]

# DMC source columns by old name
dmc_source = [
    'รหัสโรงเรียน', 'ชื่อโรงเรียน',
    'เลขประจำตัวนักเรียน',  # C in DMC = 13 digit (national_id)
    'ชั้น', 'ห้อง',
    'เลขประจำตัวนักเรียน',  # F in DMC = 4 digit — duplicate header!
    'เพศ', 'คำนำหน้าชื่อ', 'ชื่อ', 'นามสกุล',
    'วันเกิด', 'อายุ(ปี)', 'น้ำหนัก', 'ส่วนสูง', 'กลุ่มเลือด',
    'ศาสนา', 'เชื้อชาติ', 'สัญชาติ',
    'บ้านเลขที่', 'หมู่', 'ถนน/ซอย', 'ตำบล', 'อำเภอ', 'จังหวัด',
    'ชื่อผู้ปกครอง', 'นามสกุลผู้ปกครอง', 'อาชีพของผู้ปกครอง',
    'ความเกี่ยวข้องของผู้ปกครองกับนักเรียน',
    'ชื่อบิดา', 'นามสกุลบิดา', 'อาชีพของบิดา',
    'ชื่อมารดา', 'นามสกุลมารดา', 'อาชีพของมารดา',
    'ความด้อยโอกาส',
]

# Build merged rows
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = 'Students'
# write headers
for c, h in enumerate(output_headers, start=1):
    ws_out.cell(row=1, column=c, value=h)

# Read DMC raw (need to handle duplicate "เลขประจำตัวนักเรียน" header)
# Re-read DMC by column index for safety
def read_dmc_by_col_idx():
    wb = openpyxl.load_workbook(DMC_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    data = {}
    for row in range(3, ws.max_row + 1):
        sid_raw = ws.cell(row=row, column=6).value  # F
        if not sid_raw: continue
        sid = str(int(sid_raw)) if isinstance(sid_raw, float) else str(sid_raw).strip()
        if not sid: continue
        # cols by index (1-based): A=1, B=2, C=3, D=4, E=5, F=6, ...
        vals = [ws.cell(row=row, column=c).value for c in range(1, 37)]  # 36 columns A-AJ
        data[sid] = vals
    return data

def read_sheets_by_col_idx():
    wb = openpyxl.load_workbook(SHEETS_FILE, data_only=True)
    ws = wb['Student']
    data = {}
    for row in range(2, ws.max_row + 1):
        sid_raw = ws.cell(row=row, column=6).value
        if not sid_raw: continue
        sid = str(int(sid_raw)) if isinstance(sid_raw, float) else str(sid_raw).strip()
        if not sid: continue
        vals = [ws.cell(row=row, column=c).value for c in range(1, 42)]  # 41 columns A-AO
        data[sid] = vals
    return data

dmc_vals = read_dmc_by_col_idx()
sheets_vals = read_sheets_by_col_idx()

# column indices (0-based) in DMC
# A=0 รหัสโรงเรียน
# B=1 ชื่อโรงเรียน
# C=2 เลขประจำตัวประชาชน (13 หลัก)
# D=3 ชั้น
# E=4 ห้อง
# F=5 เลขประจำตัวนักเรียน (4 หลัก)
# G=6 เพศ, H=7 คำนำหน้า, I=8 ชื่อ, J=9 นามสกุล
# K=10 วันเกิด, L=11 อายุ, M=12 น้ำหนัก, N=13 ส่วนสูง, O=14 กลุ่มเลือด
# P=15 ศาสนา, Q=16 เชื้อชาติ, R=17 สัญชาติ
# S=18 บ้านเลขที่, T=19 หมู่, U=20 ถนน/ซอย, V=21 ตำบล, W=22 อำเภอ, X=23 จังหวัด
# Y=24 ชื่อผปค, Z=25 นามสกุลผปค, AA=26 อาชีพผปค, AB=27 ความเกี่ยวข้อง
# AC=28 ชื่อบิดา, AD=29 นามสกุลบิดา, AE=30 อาชีพบิดา
# AF=31 ชื่อมารดา, AG=32 นามสกุลมารดา, AH=33 อาชีพมารดา
# AI=34 ความด้อยโอกาส
# (DMC ไม่มี AJ-AO ที่เป็น GAS-added)

DMC_COLS = list(range(35))  # indices 0-34 = A-AI

stat_dmc = stat_sheets = stat_recovered = 0

for sid in all_sids:
    use_dmc = sid in dmc_vals
    src = dmc_vals.get(sid) if use_dmc else sheets_vals.get(sid)
    if not src: continue
    if use_dmc:
        stat_dmc += 1
    else:
        stat_sheets += 1

    # standard fields A-AI (35 cols)
    out_row = []
    for i in range(35):
        v = src[i] if i < len(src) else None
        # ฟื้น address (col S = index 18) จาก Date เป็น "D/M"
        if i == 18 and isinstance(v, datetime):
            v = f'{v.day}/{v.month}'
            stat_recovered += 1
        out_row.append(to_str(v))

    # status (from Sheets AK = index 36) + year (AO = index 40)
    sheet_data = sheets_vals.get(sid, [])
    status = sheet_data[36] if len(sheet_data) > 36 else None
    year = sheet_data[40] if len(sheet_data) > 40 else None
    out_row.append(to_str(status) or 'ปกติ')
    out_row.append(to_str(year))

    # write to xlsx (skip empty rows at start)
    next_row = ws_out.max_row + 1
    for c, val in enumerate(out_row, start=1):
        ws_out.cell(row=next_row, column=c, value=val)

wb_out.save(OUTPUT_FILE)
print()
print(f'✅ เซฟไฟล์ {OUTPUT_FILE}')
print(f'   📊 รวม {len(all_sids)} คน')
print(f'   📂 จาก DMC: {stat_dmc} คน (ที่อยู่สะอาด)')
print(f'   📂 จาก Sheets: {stat_sheets} คน (ใหม่ที่ DMC ไม่มี)')
print(f'   🔧 ฟื้น address จาก Date: {stat_recovered} ครั้ง')
